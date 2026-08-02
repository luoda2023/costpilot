"""
工程助手 - AI 智能辅助 API

提供 AI 驱动的智能导入/填充功能:
  POST /api/v1/ai/parse-table 理解任意格式表格数据,返回结构化字段
  POST /api/v1/ai/import-excel 上传 Excel 文件,AI 完整读取后填充
  POST /api/v1/ai/fill-fields 根据模板字段+用户描述,AI 自动填充字段值
  POST /api/v1/ai/parse-project 根据用户一句话描述,提取项目信息

设计原则:
  - 不要求用户上传固定格式文件
  - AI 理解内容后自动映射到系统字段
  - 所有接口返回结构化 JSON,前端直接填充
"""
import json
import io
import re
from typing import List, Optional, Dict, Any
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
import requests as req_lib
from packages.server.ai.client import get_ai_client, AIClientError, AIConfigError
from packages.server.ai.image_client import get_image_client, ImageClientError
from packages.server.utils.format import to_half_width
from packages.server.utils.logger import logger

router = APIRouter()

# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------

class ParseTableIn(BaseModel):
  """表格解析请求"""
  content: str  # 文件文本内容(CSV/Excel 转 text/json)
  source_type: str = "auto"  # auto / csv / json / text
  target_fields: List[str] = ["item_name", "specialty", "unit", "qty", "price"]

class ParseTableOut(BaseModel):
  """表格解析响应"""
  rows: List[Dict[str, Any]]
  total: int
  parsed: int
  column_mapping: Dict[str, str] = {}  # {field_name: original_header_text}

class FillFieldsIn(BaseModel):
    """字段填充请求"""
    fields: List[Dict[str, Any]]  # [{field_key, field_label, field_type, required, default_value}]
    description: str  # 用户描述,如"某高层住宅1#楼,建筑面积12000㎡,框架结构"

class FillFieldsOut(BaseModel):
  """字段填充响应"""
  values: Dict[str, Any]  # {field_key: filled_value}, 支持字符串/数字/布尔值

class ParseProjectIn(BaseModel):
  description: str  # 如"帮我建一个北京某高层住宅的估算项目"

class ParseProjectOut(BaseModel):
 name: str
 region: str
 stage: str
 note: str = ""

class DocSectionIn(BaseModel):
    """单节文档生成请求"""
    section_title: str  # 节标题,如"一、总论"
    section_key: str  # 节标识,如"general"
    doc_type: str  # 文档类型: bid/proposal/prelim/draw/feas/constr/contract/cost
    stage: str  # 阶段: feasibility/preliminary/construction/building/settlement
    eng_type: str  # 工程类型: pipeline/road/building/water/landscape/mep/other
    project_info: Dict[str, Any]  # 项目信息
    word_count: int = 500  # 目标字数
    context: str = ""  # 前文内容(用于保持风格一致)
    has_image: bool = False  # 是否包含配图

class DocSectionOut(BaseModel):
 """单节文档生成响应"""
 section_title: str
 section_key: str
 content: str  # Markdown 格式内容
 actual_words: int  # 实际生成字数

# ---------------------------------------------------------------------------
# 辅助函数
# ---------------------------------------------------------------------------

def _call_ai(system_prompt: str, user_prompt: str) -> str:
    """调用 AI 并返回响应文本,自动提取 JSON 块"""
    client = get_ai_client()
    resp = client.chat(messages=[
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ])
    content = resp.get("content", "")
    # 尝试提取 JSON 块(跳过 markdown 代码块标记)
    content = content.strip()
    # 先处理 markdown 代码块 ```json ... ```
    if content.startswith("```"):
        lines = content.split("\n")
        start = next((i for i, l in enumerate(lines) if "```" in l), -1)
        end = next((i for i in range(start + 1, len(lines)) if "```" in lines[i]), len(lines))
        if start >= 0:
            content = "\n".join(lines[start + 1:end]).strip()
    # 如果内容不是纯 JSON(有中文等),尝试用正则提取第一个 { ... } 或 [ ... ]
    if not (content.startswith("{") or content.startswith("[")):
        import re
        # 找 {...} 或 [...] 块
        for pattern in [r'(\{.*\})', r'(\[.*\])']:
            m = re.search(pattern, content, re.DOTALL)
            if m:
                content = m.group(1)
                break
    return content

# ---------------------------------------------------------------------------
# 路由
# ---------------------------------------------------------------------------

@router.post("/parse-table", response_model=ParseTableOut)
def parse_table(payload: ParseTableIn):
    """
    AI 理解任意格式表格数据,返回结构化字段

    流程:
    1. 接收用户上传的表格文本内容(CSV/Excel 转 text)
    2. AI 理解表格列含义,映射到目标字段
    3. 返回结构化 JSON 数组
    """
    if not payload.content.strip():
        raise HTTPException(400, "内容为空")

    system_prompt = """你是一个造价工程师,负责解析各种格式的表格数据。
表格可能来自 Excel、CSV 或文本格式,列名可能不标准(如"项目名称/材料名称/名称"→item_name)。

请理解表格的列含义,将每行数据映射到以下目标字段:
- item_name: 项目/材料名称(必填)
- specialty: 专业分类(土建/市政/安装/装饰/园林/钢结构/门窗幕墙/涂料等)
- unit: 单位(如 m³/m²/t/个/套/根/㎡/m)
- qty: 数量(数字,没有则填0)
- price: 综合单价(数字,没有则填0)

输出严格 JSON 格式:
{"rows":[{"item_name":"...","specialty":"...","unit":"...","qty":0,"price":0}]}

不要输出任何其他文字。"""

    try:
        raw = _call_ai(system_prompt, f"请解析以下表格数据:\n\n{payload.content[:10000]}")
        # 解析 JSON
        data = json.loads(raw)
        rows = data.get("rows", [])
        # 全角→半角清洗
        for row in rows:
            for key in ('item_name', 'specialty', 'unit', 'price', 'qty'):
                if key in row and isinstance(row[key], str):
                    row[key] = to_half_width(row[key])
        return ParseTableOut(
            rows=rows,
            total=len(rows),
            parsed=sum(1 for r in rows if r.get("item_name")),
            column_mapping={},
        )
    except json.JSONDecodeError:
        logger.warning("parse-table AI 返回非 JSON: %.200s", raw)
        raise HTTPException(422, "AI 返回格式异常，请重试")
    except (AIConfigError, AIClientError) as e:
        logger.error("parse-table AI 服务不可用: %s", e)
        raise HTTPException(503, "AI 服务暂不可用，请检查配置")
    except Exception as e:
        logger.error("parse-table 解析失败: %s", e, exc_info=True)
        raise HTTPException(500, "表格解析失败，请稍后重试")


@router.post("/import-excel")
async def import_excel(file: UploadFile = File(...)):
    """
    AI 读取上传的 Excel 文件,理解全部数据后返回结构化字段

    流程:
    1. 接收用户上传的 Excel/CSV 文件
    2. 用 openpyxl 完整读取全部行
    3. 本地规则解析表头列映射(优先，快速准确)
    4. 本地无法识别的列交给 AI 补充
    5. 返回结构化 JSON 数组
    """
    ALLOWED_EXTENSIONS = ('.xlsx', '.xls', '.csv')
    MAX_FILE_SIZE = 10 * 1024 * 1024
    MAX_ROWS = 500
    filename = file.filename or 'file.xlsx'
    ext = '.' + filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"不支持的文件格式: {ext}，请上传 .xlsx / .xls / .csv 文件")

    try:
        raw = await file.read()
        if len(raw) > MAX_FILE_SIZE:
            raise HTTPException(413, "文件过大，请上传 10MB 以内的文件")
    except HTTPException:
        raise
    except Exception as e:
        logger.error("import-excel 文件读取失败: %s", e, exc_info=True)
        raise HTTPException(400, "文件读取失败，请检查文件是否损坏")

    # --- 读取表格 ---
    table_rows = []
    try:
        if ext == '.csv':
            import csv
            content = raw.decode('utf-8-sig', errors='replace')
            reader = csv.reader(io.StringIO(content))
            for row in reader:
                table_rows.append(row)
        else:
            wb = openpyxl_load_workbook(raw)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                table_rows.append([str(v) if v is not None else '' for v in row])
    except Exception as e:
        logger.error("import-excel openpyxl 解析失败: %s", e, exc_info=True)
        raise HTTPException(422, "Excel 解析失败，请确保文件内容为有效表格格式")

    if len(table_rows) < 2:
        raise HTTPException(400, "文件内容不足，需要至少包含表头和一行数据")

    # --- Step 1: 本地规则解析表头列映射 ---
    headers = table_rows[0]
    data_rows = table_rows[1:]

    # 列名关键词映射表
    # 注意: 关键词按优先级排序，精确匹配优先于模糊匹配
    COLUMN_RULES = {
        'item_name': ['项目名称', '项目名', '材料名称', '材料名', '名称', '清单项目', '项目', '名称规格', '分部分项', '项目特征', '名称及规格', '构件名称', '部位名称'],
        'unit': ['计量单位', '单位(m3)', '单位(m2)', '单位(m)', '单位(t)', '单位(个)', '单位'],
        'qty': ['工程量', '数量', '工程量(m)', '工程量(m2)', '工程量(m3)', '合计数量', '设计数量', '清单工程量', '实算工程量'],
        'price': ['单价', '综合单价', '预算单价', '信息价', '市场价', '参考价', '合价'],
        'specialty': ['专业', '专业类别', '工程类别', '专业分类', '分部工程'],
        'remark': ['备注', '说明', '注释', '描述'],
    }

    def _match_column(header: str) -> str | None:
        """根据列名匹配目标字段"""
        h = header.strip().lower()
        h_no_space = re.sub(r'[\s\-_/（(）)]', '', h)
        for field, keywords in COLUMN_RULES.items():
            for kw in keywords:
                kw_clean = re.sub(r'[\s\-_/（(）)]', '', kw.lower())
                if kw_clean in h or kw_clean in h_no_space or h_no_space in kw_clean:
                    return field
        return None

    col_map = {}  # {field_name: col_index}
    for i, h in enumerate(headers):
        field = _match_column(h)
        if field:
            # 同一字段优先取第一个匹配
            if field not in col_map:
                col_map[field] = i

    logger.info("import-excel 列映射: %s", {k: f"{headers[v]}(列{v})" for k, v in col_map.items()})

    # --- Step 2: 本地解析数据行 ---
    rows = []
    for row_data in data_rows:
        if not any(v.strip() for v in row_data):
            continue  # 跳过空行
        r = {
            'item_name': '',
            'specialty': '',
            'unit': '',
            'qty': 0,
            'price': 0,
            'remark': '',
        }
        # 通过列映射填充
        for field, col_idx in col_map.items():
            if col_idx < len(row_data):
                val = row_data[col_idx].strip()
                if field in ('qty', 'price'):
                    # 提取数字
                    num_match = re.search(r'[\d,]+\.?\d*', val.replace(',', ''))
                    r[field] = float(num_match.group()) if num_match else 0
                else:
                    r[field] = to_half_width(val)

        # 全角→半角清洗
        for key in r:
            if isinstance(r[key], str):
                r[key] = to_half_width(r[key])

        rows.append(r)

    # --- Step 3: 如果本地解析不够好(缺少关键列)，用 AI 补充 ---
    parsed_count = sum(1 for r in rows if r.get('item_name'))
    # 判断是否需要 AI 补充: 没有 item_name 列 或 解析结果太少
    needs_ai = 'item_name' not in col_map or parsed_count < len(rows) * 0.5

    if needs_ai and rows:
        try:
            client = get_ai_client()
            # 构建简化表格文本（只发前50行避免超长）
            sample_rows = rows[:50]
            sample_text = "表头: " + "\t".join(headers) + "\n"
            for i, r in enumerate(sample_rows, 1):
                sample_text += f"第{i}行: {r.get('item_name','?')}\t{r.get('specialty','?')}\t{r.get('unit','?')}\t{r.get('qty',0)}\t{r.get('price',0)}\n"

            system_prompt = """你是一个造价工程师,正在补充解析Excel表格数据。
            部分列已由本地规则识别，请你补全未能识别的列。

            表格已有部分字段，如果某行 item_name 为空但表格中有项目名称，请填入。
            请输出严格JSON: {"rows":[{"item_name":"...","specialty":"...","unit":"...","qty":0,"price":0}]}"""

            raw_ai = client.chat(messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"请补充以下表格中缺失的 item_name 和价格字段:\n\n{sample_text}\n\n已识别列: {list(col_map.keys())}"},
            ])
            ai_text = raw_ai.get("content", "")
            data = json.loads(ai_text)
            ai_rows = data.get("rows", [])
            # 合并 AI 结果：只补充空白字段
            for i, ai_row in enumerate(ai_rows):
                if i < len(rows):
                    for key in ('item_name', 'specialty', 'unit', 'qty', 'price'):
                        if key in ai_row and ai_row[key] and not rows[i].get(key):
                            rows[i][key] = to_half_width(str(ai_row[key])) if isinstance(ai_row[key], str) else ai_row[key]
            parsed_count = sum(1 for r in rows if r.get('item_name'))
            logger.info("import-excel AI 补充后解析 %d/%d 行", parsed_count, len(rows))
        except (AIClientError, AIConfigError) as e:
            logger.warning("import-excel AI 补充失败(使用本地结果): %s", e)
        except Exception as e:
            logger.warning("import-excel AI 补充异常(使用本地结果): %s", e)

    # 过滤掉完全空白的行
    rows = [r for r in rows if r.get('item_name') or r.get('qty') or r.get('price')]

    # 限制行数
    if len(rows) > MAX_ROWS:
        rows = rows[:MAX_ROWS]

    return ParseTableOut(
        rows=rows,
        total=len(rows),
        parsed=sum(1 for r in rows if r.get('item_name')),
        # 返回列映射信息供前端展示
        column_mapping={k: headers[v] for k, v in col_map.items()},
    )


def openpyxl_load_workbook(raw_bytes):
    """加载 Excel 工作簿(兼容 xlsx 和 xls)"""
    import openpyxl
    buf = io.BytesIO(raw_bytes)
    try:
        return openpyxl.load_workbook(buf, data_only=True)
    except Exception:
        # openpyxl 无法读取 .xls 老格式, 尝试用 xlrd 回退
        try:
            import xlrd
            buf.seek(0)
            old_wb = xlrd.open_workbook(file_contents=raw_bytes)
            # 将 xlrd 数据转为 openpyxl 格式返回
            ws = old_wb.sheet_by_index(0)
            rows = []
            for row_idx in range(ws.nrows):
                rows.append([str(ws.cell_value(row_idx, c)) if ws.cell_value(row_idx, c) != '' else '' for c in range(ws.ncols)])
            # 返回一个简单的包装对象，模拟 openpyxl 的 active 和 iter_rows
            class FakeWorksheet:
                def __init__(self, data):
                    self._data = data
                def iter_rows(self, values_only=True):
                    for row in self._data:
                        yield row
            class FakeWorkbook:
                def __init__(self, data):
                    self.active = FakeWorksheet(data)
            return FakeWorkbook(rows)
        except ImportError:
            raise  # 重新抛出原始异常
        except Exception:
            raise  # 重新抛出原始异常

@router.post("/parse-project", response_model=ParseProjectOut)
def parse_project(payload: ParseProjectIn):
    """
    AI 根据用户一句话描述提取项目信息

    如: "帮我建一个北京某高层住宅的估算项目"
    → {"name": "北京某高层住宅估算", "region": "北京市", "stage": "估算"}
    """
    if not payload.description.strip():
        raise HTTPException(400, "描述内容为空")

    system_prompt = """你是一个造价工程师,从用户描述中提取项目信息.

输出严格 JSON 格式:
{
  "name": "项目名称(从描述中提取,长度不超过30字)",
  "region": "地区(必须是以下之一: 北京市/上海市/天津市/重庆市/广东省/浙江省/江苏省/四川省/山东省/湖北省/湖南省/福建省/河北省/河南省/安徽省/江西省,如果无法确定填'全国')",
  "stage": "项目阶段(估算/概算/预算/结算,如果无法确定填'估算')",
  "note": "备注(提取其他有用信息,如建筑面积、结构类型等)"
}

不要输出任何其他文字。"""

    try:
        raw = _call_ai(system_prompt, f"用户描述: {payload.description}")
        data = json.loads(raw)
        # 全角→半角清洗
        for key in ("name", "region", "stage", "note"):
            if key in data and isinstance(data[key], str):
                data[key] = to_half_width(data[key])
        return ParseProjectOut(
            name=data.get("name", "")[:30],
            region=data.get("region", "全国"),
            stage=data.get("stage", "估算"),
            note=data.get("note", ""),
        )
    except json.JSONDecodeError:
        logger.warning("parse-project AI 返回非 JSON: %.200s", raw)
        raise HTTPException(422, "AI 返回格式异常，请重试")
    except (AIConfigError, AIClientError) as e:
        logger.error("parse-project AI 服务不可用: %s", e)
        raise HTTPException(503, "AI 服务暂不可用，请检查配置")
    except Exception as e:
        logger.error("parse-project 解析失败: %s", e, exc_info=True)
        raise HTTPException(500, "项目解析失败，请稍后重试")

STAGE_NAME_MAP = {
    "feasibility": "可研/立项阶段",
    "preliminary": "初步设计/报批",
    "construction": "施工图/招投标",
    "building": "施工阶段",
    "settlement": "结算/审计",
}

ENG_TYPE_NAME_MAP = {
    "pipeline": "市政管网（给水/排水）",
    "road": "道路工程",
    "building": "建筑工程",
    "water": "水利工程",
    "landscape": "园林绿化",
    "mep": "机电安装",
    "other": "其他工程",
}

DOC_OUTLINES = {
    # ==================== 可研报告(发改投资规〔2023〕304号) ====================
    "feas": {
        "feasibility": [
            ("一、项目概述", "overview", 3500),
            ("二、项目建设背景和必要性", "necessity", 4000),
            ("三、项目需求分析与产出方案", "demand", 4000),
            ("四、项目选址与要素保障", "site", 3500),
            ("五、项目建设方案", "scheme", 5000),
            ("六、项目运营方案", "operation", 3500),
            ("七、项目投融资与财务方案", "finance", 4500),
            ("八、项目影响效果分析", "impact", 4000),
            ("九、项目风险管控方案", "risk", 3500),
            ("十、研究结论及建议", "conclusion", 3000),
        ],
        "preliminary": [
            ("一、工程概况", "overview", 3500),
            ("二、设计依据与范围", "design_basis", 3000),
            ("三、建设规模与内容", "scale", 3500),
            ("四、设计方案比选", "scheme_compare", 4500),
            ("五、推荐方案详细说明", "scheme_detail", 5000),
            ("六、主要工程量", "quantities", 3500),
            ("七、投资估算", "estimate", 4000),
            ("八、经济评价", "benefit", 3500),
        ],
        "default": [
            ("一、项目概述", "overview", 3500),
            ("二、建设内容与规模", "scale", 3500),
            ("三、技术方案", "scheme", 4000),
            ("四、投资估算", "estimate", 3500),
            ("五、结论与建议", "conclusion", 2500),
        ],
    },
    # ==================== 投标文件(深圳市招标范本2024) ====================
    "bid": {
        "feasibility": [
            ("一、项目管理机构与人员配置", "management", 3500),
            ("二、施工管理重点难点分析及应对措施", "difficulty", 4000),
            ("三、施工总体部署与总平面布置", "deployment", 4000),
            ("四、施工进度计划及保证措施", "schedule", 4000),
            ("五、主要分部分项工程施工技术方案", "construction_scheme", 6000),
            ("六、危大工程清单及安全管理措施", "dangerous", 4000),
            ("七、施工机械设备配置计划", "equipment", 3500),
            ("八、劳动力安排计划", "labor", 3000),
            ("九、质量安全文明施工保证措施", "quality_safety", 4000),
            ("十、绿色施工与环境保护措施", "environment", 3500),
        ],
        "preliminary": [
            ("一、项目管理机构", "management", 3000),
            ("二、施工重点难点分析", "difficulty", 3500),
            ("三、施工总体部署", "deployment", 3500),
            ("四、施工进度计划", "schedule", 3500),
            ("五、主要施工技术方案", "construction_scheme", 5000),
            ("六、质量安全保证措施", "quality_safety", 3500),
            ("七、资源配备计划", "resources", 3000),
            ("八、各项管理保证措施", "management_measures", 3500),
        ],
        "construction": [
            ("一、编制说明", "compile_note", 2500),
            ("二、工程概况", "overview", 3000),
            ("三、施工部署", "deployment", 3500),
            ("四、施工进度计划", "schedule", 3000),
            ("五、施工方案与技术措施", "construction_scheme", 5000),
            ("六、质量保证措施", "quality", 3500),
            ("七、安全管理措施", "safety", 3500),
            ("八、工期保证措施", "time_guarantee", 2500),
        ],
        "default": [
            ("一、编制说明", "compile_note", 2500),
            ("二、工程概况", "overview", 3000),
            ("三、施工方案", "construction_scheme", 4500),
            ("四、质量安全措施", "quality_safety", 3500),
            ("五、进度计划", "schedule", 2500),
        ],
    },
    # ==================== 初步设计说明(住建部深度规定2016版第3章) ====================
    "prelim": {
        "feasibility": [
            ("一、设计说明书", "design_desc", 4000),
            ("二、工程概况", "overview", 3500),
            ("三、设计依据与主要规范", "standards", 3000),
            ("四、建设规模与设计范围", "scale", 3500),
            ("五、总平面设计", "general_layout", 4000),
            ("六、建筑专业设计说明", "architecture", 4500),
            ("七、结构专业设计说明", "structure", 4500),
            ("八、给排水专业设计说明", "plumbing", 4000),
            ("九、电气专业设计说明", "electrical", 4000),
            ("十、暖通专业设计说明", "hvac", 3500),
            ("十一、主要技术经济指标", "economic", 3500),
            ("十二、概算书", "estimate", 4000),
        ],
        "preliminary": [
            ("一、设计说明书", "design_desc", 4000),
            ("二、工程概况", "overview", 3500),
            ("三、设计依据与规范", "standards", 3000),
            ("四、建设规模", "scale", 3000),
            ("五、总平面设计", "general_layout", 4000),
            ("六、各专业设计方案", "design_detail", 6000),
            ("七、主要工程数量", "quantities", 3500),
            ("八、施工组织建议", "construction_org", 3000),
            ("九、概算书", "estimate", 4000),
        ],
        "construction": [
            ("一、设计说明书", "design_desc", 3500),
            ("二、工程概况", "overview", 3000),
            ("三、设计依据", "standards", 2500),
            ("四、设计范围与内容", "scale", 3000),
            ("五、各专业设计方案", "design_detail", 5000),
            ("六、主要工程量", "quantities", 3000),
            ("七、概算", "estimate", 3500),
        ],
        "default": [
            ("一、设计说明书", "design_desc", 3500),
            ("二、工程概况", "overview", 3000),
            ("三、设计方案", "design_detail", 4500),
            ("四、主要工程量", "quantities", 2500),
            ("五、概算", "estimate", 3000),
        ],
    },
    # ==================== 施工图设计说明(住建部深度规定2016版第4章) ====================
    "draw": {
        "feasibility": [
            ("一、设计说明书", "design_desc", 3500),
            ("二、工程概况", "overview", 3000),
            ("三、设计依据与规范", "standards", 2500),
            ("四、建筑专业设计说明", "architecture", 4500),
            ("五、结构专业设计说明", "structure", 4500),
            ("六、给排水专业设计说明", "plumbing", 4000),
            ("七、电气专业设计说明", "electrical", 4000),
            ("八、施工注意事项", "construction_req", 4000),
            ("九、材料与设备表", "materials", 3500),
        ],
        "preliminary": [
            ("一、设计说明书", "design_desc", 3500),
            ("二、工程概况", "overview", 3000),
            ("三、设计依据", "standards", 2500),
            ("四、建筑专业设计", "architecture", 4500),
            ("五、结构专业设计", "structure", 4500),
            ("六、给排水专业设计", "plumbing", 4000),
            ("七、电气专业设计", "electrical", 4000),
            ("八、施工要求", "construction_req", 3500),
            ("九、材料设备清单", "materials", 3000),
        ],
        "construction": [
            ("一、设计说明书", "design_desc", 3000),
            ("二、工程概况", "overview", 2500),
            ("三、设计依据", "standards", 2500),
            ("四、各专业设计说明", "design_detail", 5000),
            ("五、施工要求", "construction_req", 4000),
            ("六、质量验收标准", "quality", 3500),
            ("七、材料设备表", "materials", 3000),
        ],
        "default": [
            ("一、设计说明", "design_desc", 3000),
            ("二、工程概况", "overview", 2500),
            ("三、各专业设计说明", "design_detail", 4500),
            ("四、施工要求", "construction_req", 3500),
            ("五、材料设备表", "materials", 2500),
        ],
    },
    # ==================== 施工组织设计(GB/T 50502-2009) ====================
    "constr": {
        "feasibility": [
            ("一、编制依据", "basis", 2500),
            ("二、工程概况", "overview", 3500),
            ("三、施工总体部署", "deployment", 4000),
            ("四、施工总进度计划", "schedule", 3500),
            ("五、施工总平面布置", "layout", 3500),
            ("六、主要施工方法", "construction_scheme", 6000),
            ("七、施工准备与资源配置计划", "resources", 3500),
            ("八、质量管理体系与措施", "quality", 4000),
            ("九、安全管理体系与措施", "safety", 4000),
            ("十、环境管理计划", "environment", 3500),
            ("十一、成本管理计划", "cost", 3000),
            ("十二、进度管理计划", "schedule_manager", 3000),
            ("十三、其他管理计划", "other", 2500),
        ],
        "preliminary": [
            ("一、编制依据", "basis", 2500),
            ("二、工程概况", "overview", 3500),
            ("三、施工部署", "deployment", 3500),
            ("四、施工进度计划", "schedule", 3500),
            ("五、施工平面布置", "layout", 3000),
            ("六、主要施工方案", "construction_scheme", 5000),
            ("七、质量保证措施", "quality", 3500),
            ("八、安全保障措施", "safety", 3500),
            ("九、资源配置", "resources", 2500),
        ],
        "construction": [
            ("一、编制依据", "basis", 2000),
            ("二、工程概况", "overview", 3000),
            ("三、施工部署", "deployment", 3500),
            ("四、施工进度计划", "schedule", 3000),
            ("五、施工方案", "construction_scheme", 5000),
            ("六、质量安全措施", "quality_safety", 3500),
            ("七、环境保护措施", "environment", 2500),
        ],
        "default": [
            ("一、编制依据", "basis", 2000),
            ("二、工程概况", "overview", 3000),
            ("三、施工部署与方案", "construction_scheme", 4500),
            ("四、质量安全保障", "quality_safety", 3500),
            ("五、进度与资源配置", "schedule_resources", 2500),
        ],
    },
    # ==================== 方案说明/比选方案 ====================
    "proposal": {
        "feasibility": [
            ("一、方案概述", "intro", 2500),
            ("二、现状分析与问题诊断", "status", 3500),
            ("三、方案比选", "compare", 4500),
            ("四、推荐方案详细说明", "scheme_detail", 5000),
            ("五、预期效果分析", "effect", 3500),
            ("六、投资估算", "estimate", 3500),
            ("七、实施计划", "implementation", 2500),
        ],
        "preliminary": [
            ("一、方案概述", "intro", 2500),
            ("二、现状分析", "status", 3000),
            ("三、方案比选", "compare", 4000),
            ("四、推荐方案", "scheme_detail", 4500),
            ("五、投资估算", "estimate", 3000),
            ("六、实施建议", "implementation", 2500),
        ],
        "construction": [
            ("一、方案概述", "intro", 2000),
            ("二、方案说明", "scheme_detail", 4500),
            ("三、投资估算", "estimate", 3000),
            ("四、实施计划", "implementation", 2500),
        ],
        "default": [
            ("一、方案概述", "intro", 2000),
            ("二、方案说明", "scheme_detail", 4000),
            ("三、投资估算", "estimate", 2500),
        ],
    },
    # ==================== 概算/目标成本(GB/T 51095) ====================
    "cost": {
        "feasibility": [
            ("一、编制说明", "compile_note", 2500),
            ("二、编制依据", "basis", 2500),
            ("三、工程概况", "overview", 2500),
            ("四、投资估算汇总表", "estimate_table", 3500),
            ("五、各专业造价分析", "cost_analysis", 4000),
            ("六、单方造价指标", "unit_cost", 2500),
            ("七、投资合理性分析", "reasonability", 2500),
        ],
        "preliminary": [
            ("一、编制说明", "compile_note", 2500),
            ("二、编制依据", "basis", 2500),
            ("三、工程概况", "overview", 2500),
            ("四、概算书", "estimate_table", 3500),
            ("五、各专业概算", "cost_analysis", 4000),
            ("六、技术经济指标", "unit_cost", 2500),
            ("七、概算对比分析", "comparison", 2500),
        ],
        "construction": [
            ("一、编制说明", "compile_note", 2000),
            ("二、编制依据", "basis", 2000),
            ("三、工程概况", "overview", 2000),
            ("四、预算书", "estimate_table", 3500),
            ("五、各专业预算", "cost_analysis", 3500),
            ("六、指标分析", "unit_cost", 2500),
        ],
        "default": [
            ("一、编制说明", "compile_note", 2000),
            ("二、编制依据", "basis", 2000),
            ("三、造价汇总", "estimate_table", 3000),
            ("四、造价分析", "cost_analysis", 3000),
        ],
    },
    # ==================== 合同范本 ====================
    "contract": {
        "feasibility": [
            ("一、合同协议书", "agreement", 3500),
            ("二、通用合同条款", "general_clauses", 4500),
            ("三、专用合同条款", "special_clauses", 4500),
            ("四、工程质量与验收", "quality", 3500),
            ("五、合同价款与支付", "payment", 4000),
            ("六、变更与索赔", "change", 3500),
            ("七、违约责任", "penalty", 2500),
            ("八、争议解决", "dispute", 2500),
        ],
        "preliminary": [
            ("一、合同协议书", "agreement", 3000),
            ("二、合同条款", "general_clauses", 4000),
            ("三、专用条款", "special_clauses", 4000),
            ("四、价款与支付", "payment", 3500),
            ("五、质量与验收", "quality", 3000),
            ("六、违约责任", "penalty", 2500),
        ],
        "construction": [
            ("一、合同主要条款", "general_clauses", 4000),
            ("二、工程范围与内容", "scope", 3000),
            ("三、价款与支付", "payment", 3500),
            ("四、质量与工期", "quality", 3000),
            ("五、违约责任", "penalty", 2500),
        ],
        "default": [
            ("一、合同主要条款", "general_clauses", 3500),
            ("二、价款与支付", "payment", 3000),
            ("三、质量与工期", "quality", 2500),
            ("四、违约责任", "penalty", 2000),
        ],
    },
}

_DOC_TYPE_ALIAS = {
    # 旧名称兼容映射
    "bid": "bid",
    "proposal": "proposal",
    "prelim": "prelim",
    "draw": "draw",
    "feas": "feas",
    "constr": "constr",
    "contract": "contract",
    "cost": "cost",
}

def _resolve_outline(doc_type: str, stage: str) -> list:
    real_type = _DOC_TYPE_ALIAS.get(doc_type, doc_type)
    type_outlines = DOC_OUTLINES.get(real_type)
    if type_outlines is None:
        type_outlines = DOC_OUTLINES.get("bid", {})
    return type_outlines.get(stage, type_outlines.get("default", []))

# 各章节详细子提纲（按章节key索引）
_SECTION_DETAIL = {
    # ===== 可研报告 =====
    "overview": "包含：项目名称、建设单位、建设地点、建设性质、建设规模、建设内容、投资总额、建设周期、编制依据(法律法规/行业规划/批准文件)、主要技术经济指标表(用地面积/建筑面积/容积率/建筑密度/绿地率/停车位/总投资/单方造价)、可行性研究结论概述",
    "necessity": "包含：项目背景(宏观政策/行业背景/区域发展现状)、建设必要性分析(需求驱动/短板补齐/政策要求)、建设的可行性条件(技术/经济/环境)、有利条件与制约因素分析、与相关规划的衔接",
    "demand": "包含：市场现状调查、供需预测分析、目标市场定位、服务需求分析、竞争力分析、市场风险分析、需求预测数据表",
    "site": "包含：选址方案比选(至少2个方案)、地理位置与交通条件、自然条件(气象/水文/地质/地形地貌)、社会经济条件、用地与规划符合性、建设条件评价、推荐选址结论",
    "scheme": "包含：技术方案比选(多方案技术经济比较表)、推荐方案详细说明、总平面布置(功能分区/交通组织/消防通道/绿化)、建筑方案(平面/立面/剖面)、结构方案(体系选型/基础形式)、给排水方案、电气方案、暖通方案、消防方案、主要设备方案(设备清单表)、土建工程方案、公用工程方案",
    "operation": "包含：运营模式选择、组织架构方案(组织架构图)、人员配置(岗位编制表)、安全保障方案、运营绩效评价体系",
    "finance": "包含：投资估算编制依据、建设投资估算(建筑工程费/设备购置费/安装工程费/其他费用/预备费分项表)、流动资金估算、资金筹措方案(资金来源/比例/使用计划)、财务评价指标表(NPV/IRR/回收期)",
    "impact": "包含：经济影响分析、社会影响分析、生态环境影响评价、环境保护措施(废气/废水/固废/噪声治理)、节能节水措施、资源能源消耗分析、碳达峰碳中和对策",
    "risk": "包含：风险因素识别表(政策/市场/技术/管理/自然)、风险程度评估(概率/影响矩阵)、风险应对措施、应急预案",
    "conclusion": "包含：推荐方案总体描述、推荐方案优缺点、主要对比方案、研究结论、存在问题与建议、下一步工作建议",
    # ===== 投标文件 =====
    "management": "包含：项目组织机构图(框架图)、项目经理部设置(各部门职责分工表)、主要管理人员配置及资历(简历表)、同类工程业绩(业绩清单表)、项目管理体系",
    "difficulty": "包含：工程特点分析、施工重点难点识别(技术难点/管理难点/协调难点)、针对性的应对措施表(难点-措施对照)、技术攻关方案、专家论证方案",
    "deployment": "包含：施工总体目标(质量/安全/工期/文明/环保目标值)、施工组织体系、施工区段划分(分区图)、施工总体流程(先地下后地上/先结构后装饰/先土建后安装)、施工流水段划分、交叉施工原则、总平面布置图",
    "schedule": "包含：总工期目标、总进度计划(横道图/网络图)、关键节点工期控制表、里程碑节点、进度保证措施(组织/技术/经济/合同)、赶工措施、进度风险分析",
    "construction_scheme": "包含：测量工程(控制网/放样)、地基与基础工程(土方/支护/降水/桩基方案及参数表)、主体结构工程(钢筋/模板/混凝土/砌体/钢结构施工工艺及参数)、屋面及防水工程、装饰装修工程(抹灰/地面/门窗/幕墙/吊顶/涂料)、安装工程(水/电/暖通/消防/智能化)、脚手架工程、季节性施工措施、各工序验收标准",
    "dangerous": "包含：危大工程识别清单(深基坑/高支模/脚手架/起重吊装)、专项方案编制要求、安全技术措施(计算书/构造要求)、监测监控方案(监测点布置/频率/报警值)、应急预案",
    "equipment": "包含：主要施工机械设备清单(名称/型号/数量/功率/进退场时间)、设备选型依据(技术经济比较)、特种设备管理措施、设备检测计划",
    "labor": "包含：各工种劳动力需求计划表(按月/按工种)、劳动力需求曲线图、劳动力来源及管理、劳务培训计划、农民工工资保障措施",
    "quality_safety": "包含：质量安全管理体系(组织架构图)、质量安全目标(量化指标)、质量安全责任制(责任分解表)、质量安全管理制度、质量控制流程(三检制/样板引路)、安全检查制度、应急预案(事故类型/响应流程)",
    "environment": "包含：文明施工管理体系、现场围挡及封闭管理、扬尘控制措施(喷雾/覆盖/冲洗)、噪声控制措施(监测/降噪)、废水处理措施(沉淀池/循环利用)、固体废弃物处理(分类/清运)、绿色施工措施(节材/节水/节能/节地)、环境保护目标指标表",
    # ===== 初步设计/施工图设计 =====
    "design_desc": "包含：工程设计依据(立项批文/规划许可/勘察报告/设计合同)、项目概况(地点/规模/功能/使用年限/耐火等级/抗震设防分类)、设计范围及分工(各专业设计范围)、设计指导思想和原则、主要技术经济指标表(总用地/总建筑面积/建筑密度/容积率/绿地率/停车位/建筑高度/层数)",
    "standards": "包含：国家现行规范标准清单(建筑/结构/给排水/电气/暖通各专业规范列表)、行业标准、地方标准、企业标准、设计采用的技术参数(抗震设防烈度/基本风压/雪压/耐火等级/防水等级/结构安全等级/设计使用年限)",
    "general_layout": "包含：设计依据及基础资料、场地概述(位置/地形/地貌/地质/周边环境)、总平面布置(功能分区/交通组织/消防通道/绿化/出入口)、竖向设计(标高/坡度/排水)、交通组织(车行/人行/停车)、主要技术经济指标表(用地面积/建筑面积/建筑密度/容积率/绿地率/停车位)、场地分析图",
    "architecture": "包含：设计依据、建筑设计概况(功能分区/平面布局/立面造型/剖面设计/各层功能表)、主要建筑材料选用(外墙/内墙/地面/天花/门窗材料表)、建筑防火设计(防火分区/疏散宽度/耐火等级/消防设施)、建筑节能设计(保温材料/节能计算/遮阳措施)、无障碍设计(坡道/电梯/卫生间)、防水设计(屋面/地下室/卫生间防水等级及做法)、建筑声学光学设计",
    "structure": "包含：设计依据、工程地质概况(地质条件/持力层/地基承载力)、结构设计参数(抗震设防烈度/基本风压/雪压/结构安全等级)、结构体系选型(框架/剪力墙/框剪/钢结构/混合结构比选及理由)、基础方案(桩基/筏板/独立基础比选)、主要结构材料及强度等级(混凝土/钢筋/钢材强度等级表)、荷载取值(恒载/活载/风载/地震作用取值表)、结构计算分析(计算模型/软件/主要计算结果)、抗震设计(抗震等级/构造措施)、结构缝设置",
    "plumbing": "包含：设计依据、给水系统设计(水源/用水量计算表/给水方式/管材/水泵参数)、排水系统设计(雨污分流/排水量计算/排水方式/管材)、消防给水系统(消火栓系统/自动喷淋系统/消防水池/消防泵参数)、热水系统(热源/供水方式/热水管材)、节水节能措施(节水器具/雨水回收/中水系统)、主要设备材料表(设备名称/规格/数量/参数)",
    "electrical": "包含：设计依据、供配电系统(负荷等级/负荷计算表/供电方案/变配电所/变压器参数)、照明系统(照度标准/灯具选型/照明控制)、防雷接地系统(防雷等级/接闪器/引下线/接地装置/等电位联结)、火灾自动报警系统(系统形式/探测器布置/联动控制)、弱电系统(综合布线/通信/安防/楼宇自控/智能化系统)、电气节能措施(节能灯具/智能控制)、主要设备材料表",
    "hvac": "包含：设计依据、室内外设计参数(温湿度/新风量标准表)、空调系统方案(系统形式/冷热源/末端设备比选)、通风系统设计(通风量计算/通风方式/风机参数)、防排烟系统设计(防烟分区/排烟量/加压送风)、采暖系统设计(热源/采暖方式/散热器/地暖)、冷热源设计(冷水机组/锅炉/热泵参数)、自动控制系统(控制策略/传感器/执行器)、节能措施(热回收/变频/分区控制)、主要设备材料表",
    "economic": "包含：主要技术经济指标(总用地面积/总建筑面积/建筑密度/容积率/绿地率/停车位/总投资/单方造价指标表)、经济评价(投资回收期/内部收益率)、主要技术经济指标对比分析",
    "estimate": "包含：编制依据、编制方法、概算书(建筑工程概算表/安装工程概算表/设备购置费表/其他费用表/预备费/总投资汇总表)、投资分析(各专业造价占比/单方造价指标/与同类项目对比)",
    "scale": "包含：建设规模方案比选(多方案对比表)、推荐建设规模及理由、建设内容明细表、各功能分区面积指标表、建设标准",
    "design_basis": "包含：项目批准文件(立项/规划/用地/环评批复)、规划条件(用地性质/容积率/建筑密度/限高)、勘察报告(地质条件/承载力/地下水)、设计合同(设计范围/阶段/深度要求)、主要规范规程(各专业规范清单)、设计范围界定(红线/设计阶段/专业分工)",
    "scheme_compare": "包含：多方案技术经济比较(至少2-3个方案对比表)、各方案优缺点分析(技术/经济/工期/施工难度)、推荐方案理由(综合评分表)、推荐方案详细参数(技术指标表)",
    "construction_req": "包含：施工准备要求(场地/技术/材料/设备)、施工工艺流程(各工序流程图)、各工序施工要点(测量/基础/主体/装修/安装)、质量验收标准(各分项工程验收标准表)、施工注意事项(安全/质量/成品保护)、成品保护要求(各专业保护措施)",
    "materials": "包含：主要材料清单(材料名称/规格型号/数量/单位/技术要求表)、设备清单(设备名称/规格/数量/参数/功率/产地)、材料规格型号详细说明、技术要求(材料标准/性能指标)、进场时间安排(材料进场计划表)",
    "quality": "包含：质量管理体系(组织架构/职责分工)、质量目标(分部分项工程合格率/优良率)、质量控制流程(材料检验/过程控制/验收程序)、质量管理制度(三检制/样板引路/质量奖罚)、各分部分项质量控制要点表、质量通病防治措施(常见问题/预防措施/治理方法)、成品保护措施、检测试验计划(检测项目/频率/标准)",
    # ===== 施工组织设计 =====
    "basis": "包含：施工合同(合同范围/工期/质量要求)、施工图纸(图纸清单/出图计划)、主要规范规程清单(国家标准/行业标准/地方标准/企业标准)、标准图集、法律法规、企业标准、地质勘察报告(地质条件/地下水/承载力)",
    "layout": "包含：布置原则(经济/合理/安全/环保)、临时设施布置(办公区/生活区/加工区/材料堆放区面积表)、施工道路布置(道路宽度/结构/排水)、临时用水用电布置(用水量计算/用电负荷计算/变压器容量)、垂直运输机械布置(塔吊/施工电梯位置/覆盖范围)、施工总平面图(各阶段平面布置图)",
    "resources": "包含：劳动力计划(各工种按月需求曲线图/表)、主要材料及构配件计划(材料名称/规格/数量/进场时间表)、主要机械设备配置计划(设备名称/型号/数量/进退场时间表)、资金使用计划(月度资金需求表)",
    "cost": "包含：成本管理目标(目标成本值)、成本管理体系(组织/制度/流程)、成本控制措施(人工/材料/机械/管理费控制)、成本核算方法、成本分析制度(月度分析/纠偏措施)",
    "schedule_manager": "包含：进度管理体系(组织/制度/流程)、进度计划编制(二级/三级进度计划)、进度跟踪与控制(周报/月报/例会)、进度纠偏措施(组织/技术/经济/合同措施)、赶工措施(资源增加/工艺优化/加班)",
    "seasonal": "包含：雨季施工措施(排水/防潮/材料保护)、冬季施工措施(保温/防冻/混凝土养护/测温)、高温季节施工措施(防暑降温/作息调整)、夜间施工措施(照明/降噪/安全)",
    "other": "包含：BIM应用方案(应用范围/模型精度/交付成果)、信息化管理方案(管理系统/数据采集/协同平台)、应急预案(事故应急救援预案/响应流程/演练计划)、工程资料管理方案(归档/分类/移交)、回访保修制度(保修期限/响应机制/回访计划)",
    "time_guarantee": "包含：工期目标分解(总工期/节点工期/里程碑)、关键线路控制(关键工作/自由时差/总时差)、资源保障措施(人力/材料/设备/资金)、组织保障措施(指挥体系/协调机制)、技术保障措施(方案优化/技术创新)、资金保障措施(资金计划/支付保障)",
    "service": "包含：保修期限及范围(各专业保修年限表)、保修响应机制(响应时间/维修流程/验收)、回访制度(回访频次/回访内容/记录归档)",
    "compile_note": "包含：编制依据(合同/图纸/规范/地质报告)、编制原则(科学/合理/经济/安全)、适用范围(工程范围/专业范围)、工程概况简介(名称/地点/规模/结构类型/工期要求)",
    "intro": "包含：方案编制背景(项目背景/编制原因)、方案编制依据(政策文件/技术标准/调研数据)、方案总体思路(技术路线/创新点)、方案主要技术经济指标(投资/工期/效益表)",
    "status": "包含：现状调查分析(现场踏勘/数据采集/问题梳理)、存在问题诊断(技术/管理/经济问题)、需求分析(功能需求/性能需求/工期需求)、改造必要性分析(安全/功能/节能/环保)",
    "compare": "包含：多方案描述(至少2-3方案详细说明)、各方案技术经济比较表(投资/工期/技术难度/运营成本/风险)、综合评分(评分表/权重/评分标准)、推荐方案及理由(综合评分最高/技术最优/经济合理)",
    "scheme_detail": "包含：推荐方案详细技术描述(工艺流程/技术参数/设备选型/材料选择)、工艺流程(工艺流程图/操作步骤/控制参数)、主要参数(技术参数表/性能指标)、设备选型(设备清单/型号/参数/数量/产地)、预期效果(技术/经济/环境效益预测表)",
    "effect": "包含：预期技术效果(性能指标/技术参数对比表)、预期经济效果(投资回报/成本节约/效益分析表)、预期社会效益(就业/民生/城市形象)、风险分析(技术风险/市场风险/政策风险/应对措施)",
    "implementation": "包含：实施步骤(分阶段实施计划表)、时间节点安排(甘特图/关键节点)、资源配置(人力/设备/材料/资金配置表)、保障措施(组织/技术/资金/制度保障)",
    "estimate_table": "包含：编制说明(编制范围/编制依据/取费标准)、编制依据(计价规范/定额/信息价/取费文件)、汇总表(建筑工程/安装工程/设备购置/其他费用/预备费/总投资汇总表)、各专业明细表(土建/装饰/安装/市政/园林各专业造价明细表)",
    "cost_analysis": "包含：各专业造价明细(土建/装饰/安装/市政各专业造价及占比)、造价构成分析(人工费/材料费/机械费/管理费/利润占比分析)、与同类项目对比分析(单方造价对比/指标对比)、造价合理性评价(偏高/偏低/合理原因分析)",
    "unit_cost": "包含：单方造价计算(总造价/建筑面积/单方造价)、单方造价指标分析(分专业单方造价表)、与行业标准对比(与定额指标/同类项目对比)、造价控制建议(优化方向/限额设计/价值工程)",
    "reasonability": "包含：投资合理性分析(各分项投资合理性判断)、与概算对比(超支/节约金额及百分比)、节约/超支原因分析(量差/价差/设计变更/签证)、优化建议(设计优化/施工优化/管理优化)",
    "comparison": "包含：概算与估算对比分析(各阶段投资对比表)、各专业概算对比(概算与估算分专业对比)、节约/超支原因(量差/价差/设计变更/签证/政策调整)、优化建议(限额设计/招标控制/过程控制)",
    "agreement": "包含：合同各方信息(发包人/承包人/监理人信息)、工程概况(名称/地点/规模/工期/质量标准)、合同工期(开工/竣工/节点工期)、质量标准(合格/优良/创优目标)、合同价款(总价/综合单价/措施费/规费/税金)、合同生效条件(签字盖章/预付款/保证金)",
    "general_clauses": "包含：一般约定(定义/解释/通知/送达/保密)、发包人义务(场地/图纸/支付/协调)、承包人义务(施工/管理/安全/质量/保修)、监理人(职责/权限/工作程序)、工程质量(标准/检验/验收/返工)、安全文明施工(安全责任/文明施工/环境保护)、合同价款与支付(预付款/进度款/结算款/质保金)、变更与索赔(变更范围/程序/估价/索赔)、竣工验收(条件/程序/资料/备案)、缺陷责任(责任期/维修/验收)、违约责任(违约情形/责任承担/赔偿)、争议解决(协商/调解/仲裁/诉讼)",
    "special_clauses": "包含：项目专用条款(针对性条款/特殊要求)、特殊技术要求(技术标准/工艺要求/材料要求)、特殊材料设备要求(进口材料/专项设备/品牌要求)、专项付款条件(节点付款/形象进度付款/比例)、违约责任特别约定(赔偿上限/违约金/罚款标准)",
    "payment": "包含：合同价款约定(固定总价/固定单价/可调价)、预付款支付(比例/金额/扣回方式)、进度款支付(申报/审核/支付周期/比例)、竣工结算(结算资料/审核流程/支付时间)、质保金(比例/返还条件/返还时间)、支付程序(申请/审核/审批/支付流程)",
    "change": "包含：变更范围(设计变更/工程洽商/现场签证)、变更程序(提出/审核/批准/实施流程)、变更估价原则(合同有约定/参照类似/协商/定额)、索赔程序(索赔意向/证据/计算/报告)、索赔时效(事件发生/通知/报告时间要求)",
    "penalty": "包含：违约情形(工期延误/质量不合格/安全事故/不支付)、违约责任承担方式(赔偿/修理/返工/解除合同)、违约赔偿标准(工期延误/质量不合格/安全事故赔偿标准表)、违约金计算(每日/每项/上限)、合同解除条件(根本违约/无法履行/协商一致)",
    "dispute": "包含：争议解决方式(协商/调解/仲裁/诉讼)、争议解决机构(仲裁委/法院)、争议解决程序(前置协商/调解/仲裁/诉讼)、适用法律(合同适用法律/管辖法律)",
    "scope": "包含：工程范围界定(承包范围/分包范围/甲供范围)、承包范围(施工总承包/专业承包/劳务分包范围)、甲方供应材料设备范围(甲供材清单/甲供设备清单)、分包范围(专业分包/劳务分包范围及资质要求)",
}

def _build_section_prompt(section_title, section_key, doc_type, stage, eng_type, project_info, word_count, context, has_image=False) -> str:
    # word_count=0 表示用户未输入，使用默认 2500
    if not word_count or word_count <= 0:
        word_count = 2500
    stage_name = STAGE_NAME_MAP.get(stage, stage)
    eng_name = ENG_TYPE_NAME_MAP.get(eng_type, eng_type)
    project_name = project_info.get("name", "本项目")
    location = project_info.get("location", "")
    scale = project_info.get("scale", "")
    scale_unit = project_info.get("scaleUnit", "")
    note = project_info.get("note", "")
    context_hint = ""
    detail_hint = _SECTION_DETAIL.get(section_key, "")
    if context:
        context_hint = "\n前文内容参考:\n" + context[:500] + "\n请保持与前文一致的写作风格和术语。\n"
    image_hint = ""
    if has_image:
        image_hint = "9. 在本章节末尾添加一张配图说明，格式为：![配图标题](配图描述，如：XXX示意图/流程图/结构图)\n"
    prompt = (
        "你是一位资深造价工程师和工程技术文档撰写专家。\n\n"
        "当前任务：撰写工程文档的单个章节。\n"
        "文档类型：" + doc_type + "\n"
        "编制阶段：" + stage_name + "\n"
        "工程类型：" + eng_name + "\n"
        "项目名称：" + project_name + "\n"
        "项目地点：" + location + "\n"
        "工程规模：" + scale + scale_unit + "\n"
        + ("补充说明：" + note + "\n" if note else "")
        + context_hint
        + "请撰写以下章节：\n【" + section_title + "】\n\n"
        + ("本章节应包含以下子项内容：\n" + detail_hint + "\n\n" if detail_hint else "")
        + "要求：\n"
        "1. 字数约 " + str(word_count) + " 字（不含标题），内容要充实深入，不能泛泛而谈\n"
        "2. 内容专业、具体、符合" + stage_name + "的编制深度要求\n"
        "3. 结合" + eng_name + "类项目的技术特点和规范要求\n"
        "4. 数据合理、术语准确、逻辑清晰\n"
        "5. 使用 Markdown 格式输出，包含多级标题(###/####)、表格、列表、加粗等\n"
        "6. 每个章节内再细分2-4个子节，每个子节用 ### 标题\n"
        "7. 重要数据和技术参数用表格呈现\n"
        "8. 引用具体的规范标准编号（如GB 50010-2010、GB 50500-2013）\n"
        + image_hint
        + "\n只输出本章节内容，不要包含其他章节。"
    )
    return prompt


class DocOutlineOut(BaseModel):
    sections: List[Dict[str, Any]]
    doc_type: str
    stage: str
    eng_type: str

@router.get("/doc-gen/outline", response_model=DocOutlineOut)
def get_doc_outline(doc_type: str, stage: str, eng_type: str = "default"):
    outline = _resolve_outline(doc_type, stage)
    sections = [
        {"title": title, "key": key, "word_count": wc}
        for title, key, wc in outline
    ]
    return DocOutlineOut(
        sections=sections,
        doc_type=doc_type,
        stage=stage,
        eng_type=eng_type,
    )

@router.post("/doc-gen/section", response_model=DocSectionOut)
def generate_doc_section(payload: DocSectionIn):
    try:
        client = get_ai_client()
        prompt = _build_section_prompt(
            payload.section_title, payload.section_key,
            payload.doc_type, payload.stage, payload.eng_type,
            payload.project_info, payload.word_count, payload.context,
            has_image=payload.has_image,
        )
        resp = client.chat(messages=[
            {"role": "system", "content": "你是一位资深造价工程师，擅长撰写各类工程技术文档。请严格按照要求输出。"},
            {"role": "user", "content": prompt},
        ])
        content = resp.get("content", "")
        if content.startswith("```"):
            content = "\n".join(content.split("\n")[1:])
        if content.endswith("```"):
            content = "\n".join(content.split("\n")[:-1])

        # 如果 has_image=True, 尝试调用图片 AI 生成配图
        if payload.has_image:
            try:
                image_client = get_image_client()
                image_desc = ""
                for line in content.split("\n"):
                    if "![配图" in line or "![" in line:
                        # 提取中括号内的描述文字
                        m = re.search(r"\[([^\]]+)\]", line)
                        if m:
                            image_desc = m.group(1)
                            break
                if not image_desc:
                    image_desc = f"{payload.section_title}示意图"
                result = image_client.generate(image_desc)
                if result.get("url"):
                    img_line = "\n\n" + f"![{image_desc}]({result['url']})"
                    content += img_line
                    logger.info("doc-gen 配图生成成功: %s", payload.section_title)
            except ImageClientError as e:
                logger.warning("doc-gen 配图生成失败(跳过): %s", e)
            except Exception as e:
                logger.warning("doc-gen 配图生成异常(跳过): %s", e)

        actual_words = len(content.replace("\n", "").replace(" ", "").replace("|", ""))
        logger.info("doc-gen 节生成: %s, 字数=%d", payload.section_title, actual_words)
        return DocSectionOut(
            section_title=payload.section_title,
            section_key=payload.section_key,
            content=content.strip(),
            actual_words=actual_words,
        )
    except (AIClientError, AIConfigError) as e:
        logger.error("doc-gen 节生成失败: %s", e)
        raise HTTPException(503, "AI 服务暂不可用: " + str(e))
    except Exception as e:
        logger.error("doc-gen 节生成异常: %s", e, exc_info=True)
        raise HTTPException(500, "文档节生成失败，请稍后重试")


class ImageGenerateIn(BaseModel):
    prompt: str
    size: str = "1024x1024"

@router.post("/generate-image")
def generate_image(payload: ImageGenerateIn):
    """通过图片 AI 生成配图, 返回图片 URL"""
    try:
        from packages.server.ai.image_client import get_image_client, ImageClientError
        client = get_image_client()
        result = client.generate(payload.prompt, payload.size)
        if result.get("url"):
            return {"ok": True, "url": result["url"], "revised_prompt": result.get("revised_prompt", "")}
        elif result.get("b64_json"):
            return {"ok": True, "b64": result["b64_json"], "revised_prompt": result.get("revised_prompt", "")}
        else:
            return {"ok": False, "msg": "图片生成失败，请检查图片 AI 配置"}
    except ImageClientError as e:
        return {"ok": False, "msg": str(e)}
    except Exception as e:
        return {"ok": False, "msg": "图片生成异常: " + str(e)}


@router.post("/image-to-image")
async def image_to_image(
    image: UploadFile = File(...),
    prompt: str = Form(...),
    size: str = Form("1024x1024"),
):
    """图生图：上传参考图片+优化提示词，AI优化改良图片"""
    try:
        # 验证文件类型
        allowed_types = {"image/jpeg", "image/png", "image/webp", "image/gif"}
        if image.content_type not in allowed_types:
            return {"ok": False, "msg": f"不支持的图片格式: {image.content_type}，仅支持 jpg/png/webp/gif"}

        # 读取文件内容
        img_bytes = await image.read()
        if len(img_bytes) > 5 * 1024 * 1024:
            return {"ok": False, "msg": "图片文件过大，请控制在5MB以内"}

        # 魔数校验：验证文件头而非content-type（防止伪造）
        _MAGIC_BYTES = {
            b'\x89PNG\r\n\x1a\n': 'png',
            b'\xff\xd8\xff': 'jpeg',
            b'GIF87a': 'gif',
            b'GIF89a': 'gif',
            b'RIFF': 'webp',
        }
        if not any(img_bytes[:len(magic)] == magic for magic in _MAGIC_BYTES):
            return {"ok": False, "msg": "图片格式校验失败，仅支持 PNG/JPG/WebP/GIF"}

        # 调用图片AI
        client = get_image_client()
        url = (client.base_url.rstrip("/") + "/images/variations")
        headers = {
            "Authorization": f"Bearer {client.api_key}",
        }

        # 构建 multipart 请求
        files = {
            "image": (image.filename or "image.png", img_bytes, image.content_type or "image/png"),
            "prompt": (None, prompt),
            "n": (None, "1"),
            "size": (None, size),
        }
        try:
            r = req_lib.post(url, files=files, headers=headers, timeout=client.timeout)
        except req_lib.RequestException as e:
            return {"ok": False, "msg": f"图生图 HTTP 请求失败: {e}"}

        if r.status_code != 200:
            # 如果不支持 variations，回落为文生图
            logger.warning("图生图API不支持(status=%d)，回落为文生图模式", r.status_code)
            result = client.generate(prompt, size)
            if result.get("url"):
                return {"ok": True, "url": result["url"], "revised_prompt": prompt, "mode": "fallback"}
            elif result.get("b64_json"):
                return {"ok": True, "b64": result["b64_json"], "revised_prompt": prompt, "mode": "fallback"}
            else:
                return {"ok": False, "msg": "图片生成失败，请检查图片 AI 配置"}

        data = r.json()
        img_data = data["data"][0]
        return {
            "ok": True,
            "url": img_data.get("url"),
            "b64_json": img_data.get("b64_json"),
            "revised_prompt": img_data.get("revised_prompt", prompt),
            "mode": "image-to-image",
        }
    except ImageClientError as e:
        return {"ok": False, "msg": str(e)}
    except Exception as e:
        return {"ok": False, "msg": "图生图异常: " + str(e)}
