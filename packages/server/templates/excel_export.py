"""
造价通 - Excel 报价表导出

按 GB50500 规范输出 6 张标准表:
  Sheet 1: 封面(项目信息)
  Sheet 2: 编制说明
  Sheet 3: 分部分项工程量清单计价表(一类)
  Sheet 4: 措施项目清单计价表(二类)
  Sheet 5: 其他项目清单计价表(三类-其他)
  Sheet 6: 规费/税金项目清单计价表(三类-规税)
  Sheet 7: 投标报价汇总表

依赖: openpyxl
"""
from pathlib import Path
from typing import List, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from packages.server.templates.pricing import QuoteResult


# ---------------------------------------------------------------------------
# 样式定义
# ---------------------------------------------------------------------------

TITLE_FONT = Font(name="宋体", size=18, bold=True)
HEADER_FONT = Font(name="宋体", size=11, bold=True)
CELL_FONT = Font(name="宋体", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")


def _set_border(ws, row, col, border=THIN_BORDER):
    cell = ws.cell(row=row, column=col)
    cell.border = border


def _style_header_row(ws, row_idx, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = HEADER_FILL


def _set_column_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# 6 张表生成函数
# ---------------------------------------------------------------------------

def _sheet_cover(wb, project_info: dict):
    """Sheet 1: 封面"""
    ws = wb.create_sheet("封面", 0)
    _set_column_widths(ws, [4, 30, 4, 30, 4])

    ws.merge_cells("B2:D2")
    ws.cell(row=2, column=2, value="工程造价报价表").font = TITLE_FONT
    ws.cell(row=2, column=2).alignment = CENTER
    ws.row_dimensions[2].height = 36

    info_rows = [
        ("项目名称", project_info.get("name", "")),
        ("项目所在地", project_info.get("region", "")),
        ("工程阶段", project_info.get("stage", "")),
        ("计税方式", project_info.get("tax_method", "一般计税")),
        ("编制人", project_info.get("compiler", "")),
        ("审核人", project_info.get("reviewer", "")),
        ("编制日期", project_info.get("date", datetime.now().strftime("%Y-%m-%d"))),
    ]
    for i, (k, v) in enumerate(info_rows, start=4):
        ws.cell(row=i, column=2, value=k).font = HEADER_FONT
        ws.cell(row=i, column=3, value=v).font = CELL_FONT
        for c in (2, 3):
            ws.cell(row=i, column=c).border = THIN_BORDER
            ws.cell(row=i, column=c).alignment = LEFT


def _sheet_note(wb, project_info: dict, quote: QuoteResult):
    """Sheet 2: 编制说明"""
    ws = wb.create_sheet("编制说明")
    _set_column_widths(ws, [80])
    ws.cell(row=1, column=1, value="编制说明").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 30

    notes = [
        f"一、本报价依据《建设工程工程量清单计价规范》GB50500-2013 编制。",
        f"二、项目所在地: {project_info.get('region', '')}。工程阶段: {project_info.get('stage', '')}。",
        f"三、计税方式: {project_info.get('tax_method', '一般计税')} ",
        f"四、综合单价 = 材料费 + 人工费 + 机械费 + 管理费(5%) + 利润(5%) + 规费 + 税金(9%/3%)",
        f"五、二类措施费以分部分项合计为基数,各项费率按地区取定;",
        f"六、三类规费以分部分项+措施为基数,税金以分部分项+措施+规费为基数;",
        f"七、本表工程总造价: ¥ {quote.grand_total:,.2f} 元",
        f"   一类(分部分项): ¥ {quote.category1:,.2f}",
        f"   二类(措施费):   ¥ {quote.category2:,.2f}",
        f"   三类(规费+税金): ¥ {quote.category3:,.2f}",
    ]
    for i, note in enumerate(notes, start=3):
        ws.cell(row=i, column=1, value=note).font = CELL_FONT
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 22


def _sheet_category1(wb, quote: QuoteResult):
    """Sheet 3: 分部分项工程量清单计价表(一类)"""
    ws = wb.create_sheet("分部分项工程量清单计价表")
    headers = ["序号", "项目名称", "专业", "单位", "工程量", "综合单价(元)", "合价(元)", "备注"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value="分部分项工程量清单计价表").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER

    # 表头
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    _style_header_row(ws, 2, len(headers))
    _set_column_widths(ws, [6, 36, 14, 8, 12, 14, 16, 24])

    # 数据行
    for idx, comp in enumerate(quote.items, start=1):
        r = 2 + idx
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=comp.item.item_name)
        ws.cell(row=r, column=3, value=comp.item.specialty or "")
        ws.cell(row=r, column=4, value=comp.item.unit)
        ws.cell(row=r, column=5, value=comp.item.qty)
        ws.cell(row=r, column=6, value=comp.item.price).number_format = "0.00"
        ws.cell(row=r, column=7, value=comp.total).number_format = "#,##0.00"
        ws.cell(row=r, column=8, value=comp.item.remark or "")
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.font = CELL_FONT
            cell.alignment = CENTER if c != 2 else LEFT

    # 合计行
    total_row = 2 + len(quote.items) + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    ws.cell(row=total_row, column=1, value="本页小计 / 一类合计").font = HEADER_FONT
    ws.cell(row=total_row, column=1).alignment = CENTER
    ws.cell(row=total_row, column=7, value=quote.category1).number_format = "#,##0.00"
    for c in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=c).border = THIN_BORDER
        ws.cell(row=total_row, column=c).fill = HEADER_FILL


def _sheet_category2(wb, quote: QuoteResult):
    """Sheet 4: 措施项目清单计价表(二类)"""
    ws = wb.create_sheet("措施项目清单计价表")
    headers = ["序号", "措施项目名称", "计算基数", "费率(%)", "金额(元)", "备注"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value="措施项目清单计价表(二类)").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER

    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    _style_header_row(ws, 2, len(headers))
    _set_column_widths(ws, [6, 32, 18, 12, 16, 24])

    for idx, (name, amount) in enumerate(quote.category2_detail.items(), start=1):
        r = 2 + idx
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value="分部分项合计")
        ws.cell(row=r, column=4, value=round(amount / quote.category1 * 100, 4) if quote.category1 else 0).number_format = "0.0000"
        ws.cell(row=r, column=5, value=amount).number_format = "#,##0.00"
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.font = CELL_FONT
            cell.alignment = CENTER if c != 2 else LEFT

    # 合计
    total_row = 2 + len(quote.category2_detail) + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    ws.cell(row=total_row, column=1, value="二类合计").font = HEADER_FONT
    ws.cell(row=total_row, column=1).alignment = CENTER
    ws.cell(row=total_row, column=5, value=quote.category2).number_format = "#,##0.00"
    for c in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=c).border = THIN_BORDER
        ws.cell(row=total_row, column=c).fill = HEADER_FILL


def _sheet_category3(wb, quote: QuoteResult):
    """Sheet 5: 其他/规费税金项目计价表(三类)"""
    ws = wb.create_sheet("规费税金项目计价表")
    headers = ["序号", "项目名称", "计算基数", "费率(%)", "金额(元)", "备注"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value="规费/税金项目计价表(三类)").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER

    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    _style_header_row(ws, 2, len(headers))
    _set_column_widths(ws, [6, 32, 22, 12, 16, 24])

    for idx, (name, amount) in enumerate(quote.category3_detail.items(), start=1):
        r = 2 + idx
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=name)
        # 简化基数说明
        if "规费" in name:
            ws.cell(row=r, column=3, value="分部分项+措施")
        else:
            ws.cell(row=r, column=3, value="分部分项+措施+规费")
        # 反算费率
        base = (quote.category1 + quote.category2) if "规费" in name else (quote.category1 + quote.category2 + quote.category3_detail.get("规费", 0))
        rate = round(amount / base * 100, 4) if base else 0
        ws.cell(row=r, column=4, value=rate).number_format = "0.0000"
        ws.cell(row=r, column=5, value=amount).number_format = "#,##0.00"
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.font = CELL_FONT
            cell.alignment = CENTER if c != 2 else LEFT

    total_row = 2 + len(quote.category3_detail) + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    ws.cell(row=total_row, column=1, value="三类合计").font = HEADER_FONT
    ws.cell(row=total_row, column=1).alignment = CENTER
    ws.cell(row=total_row, column=5, value=quote.category3).number_format = "#,##0.00"
    for c in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=c).border = THIN_BORDER
        ws.cell(row=total_row, column=c).fill = HEADER_FILL


def _sheet_summary(wb, project_info: dict, quote: QuoteResult):
    """Sheet 6: 投标报价汇总表"""
    ws = wb.create_sheet("投标报价汇总表", 0)
    headers = ["序号", "费用类别", "金额(元)", "占比(%)"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value="投标报价汇总表").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 30

    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    _style_header_row(ws, 2, len(headers))
    _set_column_widths(ws, [6, 36, 18, 12])

    rows = [
        ("一类 分部分项工程费", quote.category1),
        ("二类 措施项目费", quote.category2),
        ("三类 规费 + 税金", quote.category3),
    ]
    for idx, (name, amount) in enumerate(rows, start=1):
        r = 2 + idx
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=amount).number_format = "#,##0.00"
        pct = round(amount / quote.grand_total * 100, 2) if quote.grand_total else 0
        ws.cell(row=r, column=4, value=pct).number_format = "0.00"
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.font = CELL_FONT
            cell.alignment = CENTER if c != 2 else LEFT

    total_row = 2 + len(rows) + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    ws.cell(row=total_row, column=1, value="工程总造价").font = HEADER_FONT
    ws.cell(row=total_row, column=1).alignment = CENTER
    ws.cell(row=total_row, column=3, value=quote.grand_total).number_format = "#,##0.00"
    ws.cell(row=total_row, column=4, value=100).number_format = "0.00"
    for c in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=c).border = THIN_BORDER
        ws.cell(row=total_row, column=c).fill = HEADER_FILL
        ws.cell(row=total_row, column=c).font = HEADER_FONT


# ---------------------------------------------------------------------------
# 主导出函数
# ---------------------------------------------------------------------------

def export_quote_to_excel(
    quote: QuoteResult,
    project_info: dict,
    output_path: Optional[str] = None,
) -> str:
    """
    导出完整报价表到 Excel

    quote:        compose_quote 的返回值
    project_info: {name, region, stage, tax_method, compiler, reviewer, date}
    output_path:  指定输出路径

    返回: Excel 文件绝对路径
    """
    if output_path is None:
        output_path = f"./{project_info.get('name','未命名')}_报价表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = Workbook()
    # 删默认 sheet
    wb.remove(wb.active)

    _sheet_summary(wb, project_info, quote)   # Sheet 0
    _sheet_cover(wb, project_info)            # Sheet 1
    _sheet_note(wb, project_info, quote)      # Sheet 2
    _sheet_category1(wb, quote)               # Sheet 3
    _sheet_category2(wb, quote)               # Sheet 4
    _sheet_category3(wb, quote)               # Sheet 5

    wb.save(output_path)
    return str(Path(output_path).resolve())
