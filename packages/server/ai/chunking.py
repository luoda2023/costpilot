"""
造价通 - 文档分块器

支持格式:
  - .md / .txt / .markdown / .yml / .yaml / .csv / .json  -> 直接读文本
  - .docx                                                      -> python-docx 提文本
  - .xlsx / .xls                                               -> openpyxl 读为文本
  - .pdf                                                       -> pdfplumber 提文本 (M3 可选)

分块策略:
  - 按段落 + 字符数双重切分
  - chunk_size=800, overlap=100
"""
import re
from pathlib import Path
from typing import List, Dict, Iterator
import hashlib

from packages.server.config import get_config


def _make_id(source: str, idx: int) -> str:
    """生成稳定 ID(同源同 idx 不变)"""
    h = hashlib.md5(f"{source}::{idx}".encode("utf-8")).hexdigest()[:12]
    return f"chunk_{h}"


def extract_text(path: Path) -> str:
    """从文件提取纯文本"""
    ext = path.suffix.lower().lstrip(".")
    try:
        if ext in {"md", "txt", "markdown", "yml", "yaml", "csv", "json", "log"}:
            return path.read_text(encoding="utf-8", errors="replace")
        if ext == "docx":
            from docx import Document
            doc = Document(str(path))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        if ext in {"xlsx", "xls"}:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            lines = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    lines.append(" | ".join(cells))
                lines.append("")
            return "\n".join(lines)
        if ext == "pdf":
            try:
                import pdfplumber
                with pdfplumber.open(str(path)) as pdf:
                    return "\n\n".join(p.extract_text() or "" for p in pdf.pages)
            except ImportError:
                return ""
    except Exception as e:
        return f"[EXTRACT_ERROR: {e}]"
    return ""


def chunk_text(text: str, source: str) -> List[Dict]:
    """把文本切成块,带 overlap"""
    cfg = get_config().rag
    size = cfg.chunk_size
    overlap = cfg.chunk_overlap

    # 先按段落切,降低语义割裂
    paragraphs = re.split(r"\n\s*\n", text)
    chunks = []
    buf = ""
    idx = 0

    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        # 当前 buf 满了就 flush
        if len(buf) + len(para) > size and buf:
            chunks.append({"id": _make_id(source, idx), "text": buf, "metadata": {"source": source, "chunk_idx": idx}})
            idx += 1
            # overlap: 保留尾部
            buf = buf[-overlap:] + "\n\n" + para
        else:
            buf = (buf + "\n\n" + para) if buf else para

    if buf.strip() and len(buf) >= cfg.chunk_min_size:
        chunks.append({"id": _make_id(source, idx), "text": buf, "metadata": {"source": source, "chunk_idx": idx}})

    return chunks


def iter_chunks_from_file(path: Path) -> List[Dict]:
    """单文件 -> chunks list"""
    text = extract_text(path)
    if not text or text.startswith("[EXTRACT_ERROR"):
        return []
    return chunk_text(text, str(path))


if __name__ == "__main__":
    p = Path(r"H:\AI-model\价格信息库\01_土建\北京市_土建_综合单价库.md")
    if p.exists():
        chunks = iter_chunks_from_file(p)
        print(f"文件: {p.name}")
        print(f"块数: {len(chunks)}")
        for i, c in enumerate(chunks[:3]):
            print(f"\n--- 块 {i} (id={c['id']}) {len(c['text'])} 字符 ---")
            print(c["text"][:200])
